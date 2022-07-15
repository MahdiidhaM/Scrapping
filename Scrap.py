from os import openpty
import time
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import Workbook

wb = Workbook()
sh = wb.active
sh.title = 'mainscrap'



entery = input('Enter your search : ')
driver = webdriver.Chrome()
driver.get(f'https://www.youtube.com/results?search_query={entery}')
time.sleep(5)
driver.execute_script("window.scrollTo(0, 7000)")
mai = driver.find_elements(By.ID,'video-title')
print(mai)
for item in range(100):
    try:
        time.sleep(5)
        driver.execute_script("arguments[0].click();", mai[item])
        time.sleep(5)
        sub = driver.find_element(By.XPATH,'//*[@id="owner-sub-count"]')
        sh[f'A{item+2}'] = driver.current_url
        sh[f'B{item+2}'] = sub.text
        print(driver.current_url)
        print(sub.text)
        print('--------')
        time.sleep(5)
        # print(title.text)
        # print(subscribe.text)
        driver.back()
        time.sleep(10)
    except:
        pass

sh.cell(row=1,column=1).value = 'Subscribe'
sh.cell(row=1,column=2).value = 'Link'
wb.save('sample.xlsx')
    # driver.switch_to.window(driver.window_handles[0])
# time.sleep(60)
driver.quit()

